from openpyxl import *

from config import Constants

import configparser

def getConfigOption(key):
    cf = configparser.ConfigParser()
    cf.read(r"E:\study\UIAuto\Python\AppKeywordDriver\config\config.ini")
    value = cf.get("Excel-info", key)
    return int(value)


#读取excel文件
def getExcelFile(Path, SheetName):
    ExcelWBook = load_workbook(Path)
    ExcelSheet = ExcelWBook[SheetName]
    return ExcelSheet


# 读取Excel文件单元格数据
def getCellData(Path, SheetName, RowNum, ColNum):
    ExcelWBook = load_workbook(Path)
    ExcelSheet = ExcelWBook[SheetName]
    CellData = ExcelSheet.cell(row=RowNum, column=ColNum).value
    return CellData


# 得到一共多少行数据,去除第一行
def getRowCount(Path, SheetName):
    ExcelWBook = load_workbook(Path)
    ExcelSheet = ExcelWBook[SheetName]
    iNumber = ExcelSheet.max_row - 1
    return iNumber


# 得到测试用例开始的行号
def getRowContains(Path, SheetName, sTestCaseID, ColNum):
    iRowNum = 0
    rowCount = getRowCount(Path, SheetName)
    for iRowNum in range(rowCount):
        iRowNum = iRowNum + 1
        CellData = getCellData(Path, SheetName, iRowNum, ColNum)
        if CellData == sTestCaseID:
            return iRowNum
            break


# 计算一个测试用例有多少个步骤
def getTestStepsCount(Path, SheetName, sTestCaseID, iTestCaseStart):
    ExcelWBook = load_workbook(Path)
    for i in range(iTestCaseStart, getRowCount(Path, SheetName)):
        i = i + 1
        col = getConfigOption('Col_TestCaseID')
        CellData = getCellData(Path, SheetName, i, getConfigOption('Col_TestCaseID'))
        if sTestCaseID != CellData:
            number = i - iTestCaseStart
            return number
    #ExcelSheet = ExcelWBook[SheetName]
    #number = ExcelSheet.nrows
    #return number


# 构造一个往单元格写数据的方法，主要是用来写结果pass还是fail
def setCellData(Path, Result, RowNum, ColNum, SheetName):
    ExcelWBook = load_workbook(Path)
    ExcelSheet = ExcelWBook[SheetName]
    CellData = ExcelSheet.cell(row=RowNum, column=ColNum).value
    if isinstance(CellData, str):
        ExcelSheet.cell(row=RowNum, column=ColNum).value = Result
    ExcelSheet.save()
    ExcelSheet.close()


if __name__ == '__main__':

    path = r"E:\study\UIAuto\Python\dataEngine.xlsx"
    sheetname = 'TestSteps'
    stestcaseiD = 'AdvancedSearch_003'
    colnum = 1
    itestcasestart = 9
    # print(getConfigOption('Col_TestCaseID'))
    #print(getCellData(path, sheetname, 2, colnum))
    #print(getRowCount(path, sheetname))
    #print(getRowContains(path, sheetname, stestcaseiD, colnum))
    #print(getTestStepsCount(path, sheetname, stestcaseiD, itestcasestart))
    print(dir())


